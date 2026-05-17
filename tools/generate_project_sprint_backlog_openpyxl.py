from __future__ import annotations

from copy import copy
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

import generate_project_sprint_backlog_from_template as base


ROOT = base.ROOT
TEMPLATE_FILE = base.TEMPLATE_FILE
OUTPUT_FILE = ROOT / "docs" / "testcases" / "SCA_KTLN_Nhom09_10.ProjectSprintBacklog_openpyxl.xlsx"


def find_style_cells(ws):
    zero = yellow = green = red = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.style_id == 22 and zero is None:
                zero = cell
            elif cell.style_id == 23 and yellow is None:
                yellow = cell
            elif cell.style_id == 24 and green is None:
                green = cell
            elif cell.style_id == 25 and red is None:
                red = cell
        if zero and yellow and green and red:
            break
    return zero, yellow, green, red


def capture_style(source_cell):
    return {
        "_style": copy(source_cell._style),
    }


def set_cell_like(target, source_style, value=None):
    if isinstance(target, MergedCell):
        return
    target._style = copy(source_style["_style"])
    if value is not None:
        target.value = value
    else:
        target.value = None


def clear_row(ws, row_idx: int, date_start_col: int, date_end_col: int, zero_style_cell):
    for col_idx in range(1, date_end_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        if col_idx >= date_start_col:
            set_cell_like(cell, zero_style_cell, 0)
        else:
            cell.value = None


def set_task_row(ws, row_idx: int, task, phase_label, date_start_col: int, date_end_col: int, style_zero, style_yellow, style_green, style_red):
    if not isinstance(ws.cell(row=row_idx, column=1), MergedCell):
        ws.cell(row=row_idx, column=1).value = task.get("sprint", "") if task else ""

    label_in_b = phase_label in {"Họp kế hoạch Sprint", "Tạo Sprint Backlog", "Tạo tài liệu kiểm thử cho Sprint"}
    if label_in_b:
        if not isinstance(ws.cell(row=row_idx, column=2), MergedCell):
            ws.cell(row=row_idx, column=2).value = task["task_name"] if task else ""
        if not isinstance(ws.cell(row=row_idx, column=3), MergedCell):
            ws.cell(row=row_idx, column=3).value = None
    else:
        if not isinstance(ws.cell(row=row_idx, column=2), MergedCell):
            ws.cell(row=row_idx, column=2).value = phase_label if phase_label else ""
        if not isinstance(ws.cell(row=row_idx, column=3), MergedCell):
            ws.cell(row=row_idx, column=3).value = task["task_name"] if task else ""

    if not isinstance(ws.cell(row=row_idx, column=5), MergedCell):
        ws.cell(row=row_idx, column=5).value = task["responsible"] if task else ""
    if not isinstance(ws.cell(row=row_idx, column=7), MergedCell):
        ws.cell(row=row_idx, column=7).value = task["actual"] if task else None
    if not isinstance(ws.cell(row=row_idx, column=8), MergedCell):
        ws.cell(row=row_idx, column=8).value = task["estimate"] if task else None

    status_style = style_yellow
    if task and task["status"] == "late":
        status_style = style_red
    elif task and task["status"] == "early":
        status_style = style_green

    for offset, col_idx in enumerate(range(date_start_col, date_end_col + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = task["display_daily"][offset] if task and offset < len(task["display_daily"]) else 0
        is_finish = bool(task and offset == task["actual_end_idx"] and value)
        set_cell_like(cell, status_style if is_finish else style_zero, value)


def build_timeline_slots(start_day: date, end_day: date, slot_count: int):
    all_days = [start_day + timedelta(days=i) for i in range((end_day - start_day).days + 1)]
    if slot_count <= 0:
        return []
    if len(all_days) <= slot_count:
        return all_days

    last_index = len(all_days) - 1
    result = []
    used = set()
    for slot in range(slot_count):
        idx = round(slot * last_index / max(1, slot_count - 1))
        while idx in used and idx < last_index:
            idx += 1
        used.add(idx)
        result.append(all_days[idx])
    return result


def build_diagonal_schedule(tasks, slot_count: int):
    if slot_count <= 0:
        return

    last_finish_slot = max(0, slot_count - 1)
    task_count = len(tasks)
    if task_count == 0:
        return

    planned_end_indices = []
    for idx, _task in enumerate(tasks):
        idx = round(idx * last_finish_slot / max(1, task_count - 1))
        planned_end_indices.append(idx)

    actual_end_indices = []
    prev_actual = -1
    for idx, task in zip(planned_end_indices, tasks):
        adjusted = idx
        if task["status"] == "late":
            adjusted += 1
        elif task["status"] == "early":
            adjusted -= 1
        adjusted = max(0, min(last_finish_slot, adjusted))
        adjusted = max(prev_actual, adjusted)
        actual_end_indices.append(adjusted)
        prev_actual = adjusted

    for task, planned_end_idx, actual_end_idx in zip(tasks, planned_end_indices, actual_end_indices):
        task["planned_end_idx"] = planned_end_idx
        task["actual_end_idx"] = actual_end_idx

        display_daily = []
        for day in range(slot_count):
            display_daily.append(int(task["actual"]) if day <= actual_end_idx else 0)
        task["display_daily"] = display_daily

        task["actual_daily"] = [0] * slot_count
        if 0 <= actual_end_idx < slot_count:
            task["actual_daily"][actual_end_idx] = int(task["actual"])

        task["estimate_daily"] = [0] * slot_count
        if 0 <= planned_end_idx < slot_count:
            task["estimate_daily"][planned_end_idx] = int(task["estimate"])


def patch_sprint_sheet(ws, sprint_name: str, tasks, member_totals):
    layout = base.SPRINT_LAYOUTS[sprint_name]
    start_day, end_day = base.SPRINT_WINDOWS[sprint_name]
    date_start_col = base.col_to_index(layout["date_start_col"])
    date_end_col = base.col_to_index(layout["date_end_col"])
    total_slots = date_end_col - date_start_col + 1
    sprint_days = build_timeline_slots(start_day, end_day, total_slots)
    slot_count = len(sprint_days)
    build_diagonal_schedule(tasks, slot_count)

    zero_cell, yellow_cell, green_cell, red_cell = find_style_cells(ws)
    zero_style = capture_style(zero_cell)
    yellow_style = capture_style(yellow_cell)
    green_style = capture_style(green_cell)
    red_style = capture_style(red_cell)

    ws["C1"] = base.PROJECT_NAME
    ws["C2"] = sprint_name
    ws["C3"] = start_day
    ws["C4"] = end_day
    ws["B6"] = f"{sprint_name.upper()} REPORT"

    for idx, member in enumerate(base.TEAM_MEMBERS):
        row = layout["summary_rows"][idx]
        ws.cell(row=row, column=2).value = idx + 1
        ws.cell(row=row, column=3).value = member
        ws.cell(row=row, column=4).value = round(member_totals[member]["actual"], 1)
        ws.cell(row=row, column=5).value = round(member_totals[member]["estimate"], 1)

    total_row = layout["summary_total_row"]
    ws.cell(row=total_row, column=2).value = "Tổng"
    ws.cell(row=total_row, column=4).value = round(sum(member_totals[m]["actual"] for m in base.TEAM_MEMBERS), 1)
    ws.cell(row=total_row, column=5).value = round(sum(member_totals[m]["estimate"] for m in base.TEAM_MEMBERS), 1)

    header_row = layout["date_header_row"]
    for offset in range(total_slots):
        cell = ws.cell(row=header_row, column=date_start_col + offset)
        if offset < slot_count:
            cell.value = sprint_days[offset]
        else:
            cell.value = None

    planning = tasks[0:3]
    ui_tasks = tasks[3:3 + len(layout["ui_rows"])]
    ui_review = tasks[3 + len(layout["ui_rows"])]
    design_start = 4 + len(layout["ui_rows"])
    design_tasks = tasks[design_start:design_start + len(layout["design_rows"])]
    design_review = tasks[design_start + len(layout["design_rows"])]
    coding_start = design_start + len(layout["design_rows"]) + 1
    coding_tasks = tasks[coding_start:coding_start + len(layout["coding_pairs"]) * 2]
    integrate = tasks[coding_start + len(layout["coding_pairs"]) * 2]
    testing_start = coding_start + len(layout["coding_pairs"]) * 2 + 1
    testing_tasks = tasks[testing_start:testing_start + len(layout["testing_rows"])]
    fix_start = testing_start + len(layout["testing_rows"])
    fix_tasks = tasks[fix_start:fix_start + len(layout["fix_rows"])]
    retest_start = fix_start + len(layout["fix_rows"])
    retest_tasks = tasks[retest_start:retest_start + len(layout["retest_rows"])]
    release_tasks = tasks[retest_start + len(layout["retest_rows"]):retest_start + len(layout["retest_rows"]) + 2]

    used_rows = set()
    for idx, row in enumerate(layout["planning_rows"]):
        label = "Họp kế hoạch Sprint" if idx == 0 else ("Tạo Sprint Backlog" if idx == 1 else "Tạo tài liệu kiểm thử cho Sprint")
        set_task_row(ws, row, planning[idx], label, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        used_rows.add(row)
    for idx, row in enumerate(layout["ui_rows"]):
        set_task_row(ws, row, ui_tasks[idx], "User interface design" if idx == 0 else None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        used_rows.add(row)
    set_task_row(ws, layout["ui_review_row"], ui_review, None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
    used_rows.add(layout["ui_review_row"])
    for idx, row in enumerate(layout["design_rows"]):
        set_task_row(ws, row, design_tasks[idx], "Design test case" if idx == 0 else None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        used_rows.add(row)
    set_task_row(ws, layout["design_review_row"], design_review, None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
    used_rows.add(layout["design_review_row"])
    for idx, (front_row, back_row) in enumerate(layout["coding_pairs"]):
        set_task_row(ws, front_row, coding_tasks[idx * 2], "Coding" if idx == 0 else None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        set_task_row(ws, back_row, coding_tasks[idx * 2 + 1], None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        used_rows.add(front_row)
        used_rows.add(back_row)
    set_task_row(ws, layout["integrate_row"], integrate, None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
    used_rows.add(layout["integrate_row"])
    for idx, row in enumerate(layout["testing_rows"]):
        set_task_row(ws, row, testing_tasks[idx], "Testing" if idx == 0 else None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        used_rows.add(row)
    for idx, row in enumerate(layout["fix_rows"]):
        set_task_row(ws, row, fix_tasks[idx], "Fix Bug" if idx == 0 else None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        used_rows.add(row)
    for idx, row in enumerate(layout["retest_rows"]):
        set_task_row(ws, row, retest_tasks[idx], "Re-testing" if idx == 0 else None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
        used_rows.add(row)
    set_task_row(ws, layout["release_rows"][0], release_tasks[0], f"Release {sprint_name}", date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
    set_task_row(ws, layout["release_rows"][1], release_tasks[1], None, date_start_col, date_end_col, zero_style, yellow_style, green_style, red_style)
    used_rows.update(layout["release_rows"])

    for row_idx in range(layout["planning_rows"][0], layout["total_actual_row"]):
        if row_idx not in used_rows:
            clear_row(ws, row_idx, date_start_col, date_end_col, zero_style)

    total_actual_row = layout["total_actual_row"]
    total_estimate_row = layout["total_estimate_row"]
    ws.cell(row=total_actual_row, column=2).value = "Tổng"
    ws.cell(row=total_actual_row, column=5).value = "Thực tế"
    ws.cell(row=total_estimate_row, column=5).value = "Ước tính"
    total_actual = sum(int(task["actual"]) for task in tasks)
    total_estimate = sum(int(task["estimate"]) for task in tasks)
    ws.cell(row=total_actual_row, column=7).value = total_actual
    ws.cell(row=total_estimate_row, column=7).value = total_estimate
    if not isinstance(ws.cell(row=total_actual_row, column=8), MergedCell):
        ws.cell(row=total_actual_row, column=8).value = None
    if not isinstance(ws.cell(row=total_estimate_row, column=8), MergedCell):
        ws.cell(row=total_estimate_row, column=8).value = None

    actual_remaining = []
    estimate_remaining = []
    for day in range(slot_count):
        actual_remaining.append(sum(int(task["actual"]) for task in tasks if task["actual_end_idx"] >= day))
        estimate_remaining.append(sum(int(task["estimate"]) for task in tasks if task["planned_end_idx"] >= day))

    for offset in range(total_slots):
        actual_cell = ws.cell(row=total_actual_row, column=date_start_col + offset)
        estimate_cell = ws.cell(row=total_estimate_row, column=date_start_col + offset)
        if offset < slot_count:
            actual_cell.value = actual_remaining[offset]
            estimate_cell.value = estimate_remaining[offset]
        else:
            actual_cell.value = None
            estimate_cell.value = None


def patch_total_sheet(ws, sprint_task_sets):
    member_totals_by_sprint = {sprint: base.compute_member_totals(tasks) for sprint, tasks in sprint_task_sets.items()}
    for col, member in zip(["B", "D", "F", "H", "J"], base.TEAM_MEMBERS):
        ws[col + "2"] = member
    sprint_rows = {"Sprint 1": 4, "Sprint 2": 5, "Sprint 3": 6}
    for sprint_name, row in sprint_rows.items():
        ws[f"A{row}"] = sprint_name
        totals = member_totals_by_sprint[sprint_name]
        for idx, member in enumerate(base.TEAM_MEMBERS):
            ws.cell(row=row, column=2 + idx * 2).value = round(totals[member]["actual"], 1)
            ws.cell(row=row, column=3 + idx * 2).value = round(totals[member]["estimate"], 1)
    ws["A7"] = "Tổng"
    for idx, member in enumerate(base.TEAM_MEMBERS):
        ws.cell(row=7, column=2 + idx * 2).value = round(sum(member_totals_by_sprint[s][member]["actual"] for s in sprint_rows), 1)
        ws.cell(row=7, column=3 + idx * 2).value = round(sum(member_totals_by_sprint[s][member]["estimate"] for s in sprint_rows), 1)
    ws["F11"] = round(sum(sum(member_totals_by_sprint[s][m]["actual"] for m in base.TEAM_MEMBERS) for s in sprint_rows), 1)
    ws["F12"] = round(sum(sum(member_totals_by_sprint[s][m]["estimate"] for m in base.TEAM_MEMBERS) for s in sprint_rows), 1)


def generate():
    reports = [base.parse_testcase_workbook(path) for path in base.SOURCE_FILES]
    report_map = {report["sprint_name"]: report for report in reports}
    sprint_task_sets = {}
    sprint_member_totals = {}
    for sprint_name in ["Sprint 1", "Sprint 2", "Sprint 3"]:
        groups = base.build_groups(sprint_name, report_map[sprint_name]["modules"])
        tasks = base.build_sprint_tasks(sprint_name, groups)
        sprint_task_sets[sprint_name] = tasks
        sprint_member_totals[sprint_name] = base.compute_member_totals(tasks)

    wb = load_workbook(TEMPLATE_FILE)
    for sprint_name in ["Sprint 1", "Sprint 2", "Sprint 3"]:
        patch_sprint_sheet(wb[sprint_name], sprint_name, sprint_task_sets[sprint_name], sprint_member_totals[sprint_name])
    patch_total_sheet(wb["Total"], sprint_task_sets)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


if __name__ == "__main__":
    print(generate())
